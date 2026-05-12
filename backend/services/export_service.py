"""
数据导出服务
============
支持两种导出格式：
- CSV: 轻量、纯 Python 标准库，适合机器处理
- Excel: 带样式 + 图表，适合给人看的报告（需要 openpyxl）

可导出内容：
- 任务的全部性能采样点（IOPerformanceData）
- 任务的 iostat 采样（IOStatData）
- 多任务对比（横向对比平均值/峰值）
- 基准对比报告

使用：
    csv_bytes = ExportService.task_metrics_csv(db, task_id)
    xlsx_bytes = ExportService.task_report_xlsx(db, task_id)
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import List, Optional

from sqlalchemy.orm import Session

from models.task import Task, TaskNode, IOPerformanceData, IOStatData


class ExportService:

    # ---------------------------------------------------------- CSV 导出 ----

    @staticmethod
    def task_metrics_csv(db: Session, task_id: int) -> bytes:
        """
        导出某任务的全部性能采样点 CSV
        列：timestamp, node_id, node_name, iops, bandwidth_mbs, latency_ms,
             read_iops, write_iops, read_bw_mbs, write_bw_mbs, read_lat_ms, write_lat_ms,
             cpu_usage, memory_usage
        """
        rows = (
            db.query(IOPerformanceData)
            .join(TaskNode, IOPerformanceData.task_node_id == TaskNode.id)
            .filter(TaskNode.task_id == task_id)
            .order_by(IOPerformanceData.timestamp.asc())
            .all()
        )

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            "timestamp", "node_id", "node_name",
            "iops", "bandwidth_mbs", "latency_ms",
            "read_iops", "write_iops",
            "read_bw_mbs", "write_bw_mbs",
            "read_lat_ms", "write_lat_ms",
            "cpu_usage_pct", "memory_usage_pct",
        ])
        for r in rows:
            tn = r.task_node
            node = tn.node if tn else None
            writer.writerow([
                r.timestamp.isoformat() if r.timestamp else "",
                node.id if node else "",
                node.node_name if node else "",
                float(r.iops or 0),
                float(r.bandwidth or 0),
                float(r.latency or 0),
                float(r.read_iops or 0),
                float(r.write_iops or 0),
                float(r.read_bw or 0),
                float(r.write_bw or 0),
                float(r.read_lat or 0),
                float(r.write_lat or 0),
                float(r.cpu_usage or 0),
                float(r.memory_usage or 0),
            ])

        return buf.getvalue().encode("utf-8-sig")   # BOM 方便 Excel 直接打开不乱码

    @staticmethod
    def task_iostat_csv(db: Session, task_id: int) -> bytes:
        rows = (
            db.query(IOStatData)
            .join(TaskNode, IOStatData.task_node_id == TaskNode.id)
            .filter(TaskNode.task_id == task_id)
            .order_by(IOStatData.timestamp.asc())
            .all()
        )
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow([
            "timestamp", "node_id", "device", "tps",
            "kB_read_s", "kB_wrtn_s", "await_ms", "aqu_sz", "util_pct",
        ])
        for r in rows:
            tn = r.task_node
            node = tn.node if tn else None
            w.writerow([
                r.timestamp.isoformat() if r.timestamp else "",
                node.id if node else "",
                r.device or "",
                float(r.tps or 0),
                float(r.kB_read_s or 0),
                float(r.kB_wrtn_s or 0),
                float(r.await_time or 0),
                float(r.aqu_sz or 0),
                float(r.util or 0),
            ])
        return buf.getvalue().encode("utf-8-sig")

    # ---------------------------------------------------------- XLSX 报告 ----

    @staticmethod
    def task_report_xlsx(db: Session, task_id: int) -> Optional[bytes]:
        """
        生成多 sheet 的 Excel 报告：
        - Summary: 任务基本信息 + 平均值
        - Nodes: 每个节点的汇总
        - Metrics: 全部性能采样点
        - IOStat: iostat 采样
        返回 xlsx 字节流；如果 openpyxl 未安装返回 None
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.chart import LineChart, Reference
        except ImportError:
            return None

        task = db.query(Task).filter(Task.id == task_id).first()
        if not task:
            return None

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1E40AF")

        # -- Summary --
        ws = wb.active
        ws.title = "Summary"
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 40

        rows = [
            ("Task ID", task.id),
            ("Task Name", task.task_name),
            ("Status", task.status),
            ("Created At", str(task.created_at) if task.created_at else ""),
            ("Start Time", str(task.start_time) if task.start_time else ""),
            ("End Time", str(task.end_time) if task.end_time else ""),
            ("Duration (s)", task.duration or 0),
            ("Total IO Ops", task.total_io_ops or 0),
            ("Avg IOPS", float(task.avg_iops or 0)),
            ("Avg Latency (ms)", float(task.avg_latency or 0)),
            ("Avg Bandwidth (MB/s)", float(task.avg_bw or 0)),
            ("Test Case", task.test_case.case_name if task.test_case else ""),
            ("Test Case RW Mode", task.test_case.rw_mode if task.test_case else ""),
            ("Block Size", task.test_case.block_size if task.test_case else ""),
            ("IO Engine", task.test_case.io_engine if task.test_case else ""),
            ("Queue Depth", task.test_case.queue_depth if task.test_case else ""),
            ("IO Size", task.test_case.io_size if task.test_case else ""),
            ("Runtime (s)", task.test_case.runtime if task.test_case else ""),
        ]
        for idx, (k, v) in enumerate(rows, start=1):
            ws.cell(row=idx, column=1, value=k).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=v)

        # -- Nodes --
        ws2 = wb.create_sheet("Nodes")
        headers2 = [
            "node_id", "node_name", "host", "partition", "status",
            "iops", "bandwidth_mbs", "latency_ms", "io_ops",
            "duration_s", "error_message",
        ]
        for c, h in enumerate(headers2, start=1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for i, tn in enumerate(task.task_nodes or [], start=2):
            ws2.cell(row=i, column=1, value=tn.node.id if tn.node else "")
            ws2.cell(row=i, column=2, value=tn.node.node_name if tn.node else "")
            ws2.cell(row=i, column=3, value=tn.node.host if tn.node else "")
            ws2.cell(row=i, column=4, value=tn.partition.mount_point if tn.partition else "")
            ws2.cell(row=i, column=5, value=tn.status)
            ws2.cell(row=i, column=6, value=float(tn.iops or 0))
            ws2.cell(row=i, column=7, value=float(tn.bandwidth or 0))
            ws2.cell(row=i, column=8, value=float(tn.latency or 0))
            ws2.cell(row=i, column=9, value=int(tn.io_ops or 0))
            ws2.cell(row=i, column=10, value=tn.duration or 0)
            ws2.cell(row=i, column=11, value=tn.error_message or "")

        # -- Metrics （带图表）--
        ws3 = wb.create_sheet("Metrics")
        m_headers = [
            "timestamp", "node_name", "iops", "bandwidth_mbs", "latency_ms",
            "read_iops", "write_iops", "cpu_usage_pct", "memory_usage_pct",
        ]
        for c, h in enumerate(m_headers, start=1):
            cell = ws3.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill

        metrics_rows = (
            db.query(IOPerformanceData)
            .join(TaskNode, IOPerformanceData.task_node_id == TaskNode.id)
            .filter(TaskNode.task_id == task_id)
            .order_by(IOPerformanceData.timestamp.asc())
            .all()
        )
        for i, r in enumerate(metrics_rows, start=2):
            tn = r.task_node
            node = tn.node if tn else None
            ws3.cell(row=i, column=1, value=r.timestamp.strftime("%H:%M:%S") if r.timestamp else "")
            ws3.cell(row=i, column=2, value=node.node_name if node else "")
            ws3.cell(row=i, column=3, value=float(r.iops or 0))
            ws3.cell(row=i, column=4, value=float(r.bandwidth or 0))
            ws3.cell(row=i, column=5, value=float(r.latency or 0))
            ws3.cell(row=i, column=6, value=float(r.read_iops or 0))
            ws3.cell(row=i, column=7, value=float(r.write_iops or 0))
            ws3.cell(row=i, column=8, value=float(r.cpu_usage or 0))
            ws3.cell(row=i, column=9, value=float(r.memory_usage or 0))

        # 给 Metrics 加一张 IOPS 趋势折线图
        if len(metrics_rows) > 1:
            chart = LineChart()
            chart.title = "IOPS Trend"
            chart.y_axis.title = "IOPS"
            chart.x_axis.title = "Time"
            data_ref = Reference(ws3, min_col=3, max_col=3,
                                 min_row=1, max_row=len(metrics_rows) + 1)
            cats_ref = Reference(ws3, min_col=1, max_col=1,
                                 min_row=2, max_row=len(metrics_rows) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.height = 8
            chart.width = 20
            ws3.add_chart(chart, "K2")

        # -- IOStat --
        ws4 = wb.create_sheet("IOStat")
        s_headers = ["timestamp", "node_name", "device", "tps",
                     "kB_read_s", "kB_wrtn_s", "await_ms", "util_pct"]
        for c, h in enumerate(s_headers, start=1):
            cell = ws4.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
        iostat_rows = (
            db.query(IOStatData)
            .join(TaskNode, IOStatData.task_node_id == TaskNode.id)
            .filter(TaskNode.task_id == task_id)
            .order_by(IOStatData.timestamp.asc())
            .all()
        )
        for i, r in enumerate(iostat_rows, start=2):
            tn = r.task_node
            node = tn.node if tn else None
            ws4.cell(row=i, column=1, value=r.timestamp.strftime("%H:%M:%S") if r.timestamp else "")
            ws4.cell(row=i, column=2, value=node.node_name if node else "")
            ws4.cell(row=i, column=3, value=r.device or "")
            ws4.cell(row=i, column=4, value=float(r.tps or 0))
            ws4.cell(row=i, column=5, value=float(r.kB_read_s or 0))
            ws4.cell(row=i, column=6, value=float(r.kB_wrtn_s or 0))
            ws4.cell(row=i, column=7, value=float(r.await_time or 0))
            ws4.cell(row=i, column=8, value=float(r.util or 0))

        # 输出
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    # -------------------------------------------------- 多任务对比 CSV ----

    @staticmethod
    def compare_tasks_csv(db: Session, task_ids: List[int]) -> bytes:
        """
        横向对比多个任务的关键指标
        列：task_id, task_name, case_name, avg_iops, avg_latency, avg_bw, duration, status
        """
        tasks = db.query(Task).filter(Task.id.in_(task_ids)).all()
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow([
            "task_id", "task_name", "case_name",
            "avg_iops", "avg_latency_ms", "avg_bw_mbs",
            "duration_s", "status", "end_time",
        ])
        for t in tasks:
            w.writerow([
                t.id, t.task_name,
                t.test_case.case_name if t.test_case else "",
                float(t.avg_iops or 0),
                float(t.avg_latency or 0),
                float(t.avg_bw or 0),
                t.duration or 0,
                t.status,
                t.end_time.isoformat() if t.end_time else "",
            ])
        return buf.getvalue().encode("utf-8-sig")
